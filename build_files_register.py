import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    'hdr_dr':   '1565C0', 'sec_dr':   'E3F2FD',
    'hdr_dh':   '2E7D32', 'sec_dh':   'E8F5E9',
    'hdr_dps':  '6A1B9A', 'sec_dps':  'F3E5F5',
    'hdr_heit': 'E65100', 'sec_heit': 'FFF3E0',
    'hdr_sum':  '37474F', 'sec_sum':  'ECEFF1',
    'white':    'FFFFFF', 'light':    'F5F7FA',
    'border':   'BDBDBD',
    'pbix':     'E8EAF6', 'xlsx_c':   'DCEEFB',
    'xls_c':    'FFF9C4', 'xlsm_c':   'E8F5E9',
    'docx_c':   'F3E5F5',
    't_pbix':   '283593', 't_xlsx':   '0D47A1',
    't_xls':    'F57F17', 't_xlsm':   '1B5E20',
    't_docx':   '4A148C',
}

def fill(h):
    return PatternFill('solid', fgColor=h)

def fnt(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')

def aln(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr():
    s = Side(style='thin', color=C['border'])
    return Border(left=s, right=s, top=s, bottom=s)

def wc(ws, r, c, val, bold=False, fc=None, tc='000000', sz=10, ha='left', wrap=True, italic=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = fnt(bold=bold, color=tc, size=sz, italic=italic)
    cell.alignment = aln(h=ha, wrap=wrap)
    if fc:
        cell.fill = fill(fc)
    cell.border = bdr()
    return cell

TYPE_STYLE = {
    'pbix':  (C['pbix'],   C['t_pbix'],  'Power BI Report (.pbix)'),
    'xlsx':  (C['xlsx_c'], C['t_xlsx'],  'Excel Workbook (.xlsx)'),
    'xls':   (C['xls_c'],  C['t_xls'],  'Excel Legacy SAP BW (.xls)'),
    'xlsm':  (C['xlsm_c'], C['t_xlsm'], 'Excel Macro-Enabled (.xlsm)'),
    'docx':  (C['docx_c'], C['t_docx'], 'Word Document (.docx)'),
}

# ── DATA ───────────────────────────────────────────────────────────────────────
# (dashboard, file_name, file_type, sheet_count, sheet_names, role, location)

DATA = [
    # ── Digital Radiology ──
    ('Digital Radiology', 'BW data in prize realisation.xlsx', 'xlsx', 2,
     'BP5 // BP2',
     'SAP BW query spec for price realization extract (AP5 & AP2 systems, feeds Report 3)',
     'Data/'),

    ('Digital Radiology', 'CRM data in Radiology reporting.xlsx', 'xlsx', 2,
     'CRM extracts // region mapping',
     'D365 CRM field mapping (48 opp fields, 100+ product fields) + 271-row country-to-region master',
     'Data/'),

    ('Digital Radiology', 'Sales playbook Agfa Digital Radiology.docx', 'docx', 0,
     '—',
     'Sales process documentation: stages, DS%/DH%/Feasibility%, forecast categories, review cadence, large deal rules',
     'Data/'),

    ('Digital Radiology', 'Commercial Analytics - OIT Margin & Product Mix 2026.pbix', 'pbix', 8,
     'BUDGET - OIT MARGIN AND PRODUCTS MIX // TARGET - OIT MARGIN AND PRODUCTS MIX // DETAILED FUNNEL VIEW // PRODUCT PERFORMANCE // ToolTip // Intercontinental // Page 1 // Page 2',
     'Report 1 — OIT vs budget/target, product mix, margin, detailed funnel by region & equipment',
     'Power BI Reports/Dashboards1/'),

    ('Digital Radiology', 'Commercial Analytics - Weekly FC Tracker.pbix', 'pbix', 12,
     'Forecast Flags // Global FC // Region FC // SubRegion FC // Weighted Funnel Evolution // Unweighted Funnel Evolution // KAM FC // KAM Funnel // FC Chart // Opp List // New Opportunities // RawData',
     'Report 2 — Weekly forecast tracker, KAM scorecard, funnel evolution (weighted & unweighted)',
     'Power BI Reports/Dashboards2/'),

    ('Digital Radiology', 'Partner Dashboard.pbix', 'pbix', 12,
     'SAP channel check // Indirect Target CY // Sales YTD CY vs PY // Sales FY PY vs PY-1 // Sales Evolution // Partner Performance YTD CY vs PY // Partner Performance FY PY vs PY-1 // Partner Performance Evolution // Top 5 Partners // Margin // Remarks // RawData',
     'Report 3 — Partner/dealer revenue vs target, SAP actuals by channel, margin waterfall',
     'Power BI Reports/Dashboards2/'),

    ('Digital Radiology', 'Price Margin Modalities.pbix', 'pbix', 4,
     'Overview // Regions // Goods Waterfall // Raw Data',
     'Report 4 — Price realization waterfall (List Price → ENP → Gross Margin), Sofon standard margin',
     'Power BI Reports/Dashboards2/'),

    ('Digital Radiology', 'Commercial Analytics - Funnel Evolution Tracker.pbix', 'pbix', 4,
     'Unweighted // Weighted - Monthly // Unweighted - Monthly // Weighted',
     'Report 5 — Funnel evolution tracker (weighted / unweighted, weekly / monthly views)',
     'Power BI Reports/Dashboards3/'),

    ('Digital Radiology', 'Commercial Analytics - OI & Funnel Health Cockpit.pbix', 'pbix', 3,
     'Order Intake YTD // Order Intake Q // Order Intake FY',
     'Report 6 — OI & funnel health cockpit with RT CY/BT/PY running totals (YTD / Q / FY)',
     'Power BI Reports/Dashboards3/'),

    # ── Digital Hydrogen ──
    ('Digital Hydrogen', 'AP1 SAP extract on sales orders 2026.xlsx', 'xlsx', 1,
     'Blad1',
     'SAP AP1 order extract — 46 orders YTD 2026, delivery dates, confirmed quantities, order numbers',
     'Data/'),

    ('Digital Hydrogen', 'Master data _Customers overview Zirfon.xlsx', 'xlsx', 1,
     'Customer info',
     'Customer master — 304 customers, regions, sub-regions, payment terms, IC flag, address',
     'Data/'),

    ('Digital Hydrogen', 'Overview quotes GHS.xlsx', 'xlsx', 5,
     '5657-8081 // Pricing // 5638 // pivot // Sheet2',
     'Quotation pipeline — 1,337 quotes 2023-2025, EUR/m2 pricing, conversion pivot by customer/product',
     'Data/'),

    ('Digital Hydrogen', 'Sales zirfon GHS.xlsx', 'xlsx', 6,
     '2023 // 2024 // 2025 // 2026 // Stock overview per 30.09 // Forecast info',
     'Central sales hub — 920 orders across 4 annual sheets, stock overview, forecast commentary',
     'Data/'),

    ('Digital Hydrogen', 'ACTFY2025_Forecasting file.xlsx', 'xlsx', 44,
     'FOR Summary // 18 months FOR // FOR vs BUD & LY // FOR vs previous FOR // top customers & volumes // ACT data structured 2025 // ACT data structured // ACT 2023 // ACT2024 Summary // Committed volumes by Customer // TK Nucera // open orders pivot // BUD2024 // BUD2025 // all orders 2023 // month mapping // customer & country list // Calculated fields // … (44 total)',
     'FY2025 forecasting workbook — actuals, budget, 18-month rolling forecast, committed/uncommitted, TK Nucera batches',
     'Data/Sales & Margin Analyses/'),

    ('Digital Hydrogen', 'FY 2025.xls', 'xls', 26,
     'BExRepositorySheet // Summary // Month TO // Month Margin // Quarter // YTD TO // Q1 margin // Full Pivot Margin // TO pivot // BI Report Mortsel // BI Report Aspac // BI Report Turnover Analysis // Mapping Standard Costprices // Mapping Customers // Mapping Company Code // Mapping Type // Mapping Months // raw data // Summary Views // sheet to be used for FOR // … (26 total)',
     'SAP BW controller workbook — turnover & margin by month/quarter, standard cost prices, budget vs LY, BI reports',
     'Data/Sales & Margin Analyses/'),

    ('Digital Hydrogen', 'Sales Forecast February2026.xlsx', 'xlsx', 49,
     'FOR Summary // Revenue overview // 18 months FOR // FOR vs BUD & LY // FOR vs previous FOR // March FOR // FOR_TO_QTY_ENP // top customers & volumes // ACT data structured 2026 // ACT data structured 2025 // BUD2026 // BUD2025 // open orders 2026 // Customer sort 2026 // TK Nucera // Committed volumes by Customer // Calculated fields // … (49 total)',
     'Latest FY2026 forecast — updated actuals, BUD2026, revenue overview, March FOR, Customer sort 2026',
     'Data/Sales & Margin Analyses/'),

    # ── DPS ──
    ('DPS', 'DPS_Customer order & revenue follow-up 2026.xlsx', 'xlsx', 62,
     'SPR printer OIT 2026 // SPR Tauro periph OIT 2026 // Delayed // RR FULL YEAR 2023 // RR FULL YEAR 2024 // RR 01-12_2025 Eq Fam (12 sheets) // RR 01-02_2026 Eq Fam // Potential/Expected RR 03-05_2026 // OIT FY 2023-2025 // actual OIT 01-04 2026 // Dashboard 2026 // SSOT RR 2025 // RR 2025/2026 YTD // RR 2026 YTD FAMILY // Expected RR 2026 Cntry // AD // EFI assortm // … (62 total)',
     'Master order register — 2,501 rows, monthly OIT/RR snapshots by equipment family, delayed tracker, SSOT, Dashboard 2026',
     'Data/Order-Intake/'),

    ('DPS', 'DPS Equipment Customer order & revenue follow-up 2026 for Sutherland project.xlsx', 'xlsx', 61,
     'SPR printer OIT 2026 // SPR Tauro periph OIT 2026 // Delayed // RR FULL YEAR 2023/2024 // RR 01-12_2025 Eq Fam // RR 01-02_2026 Eq Fam // Potential/Expected RR 03-05_2026 // OIT monthly 2025-2026 // Dashboard 2026 // SSOT // Printer Overview 2024/2025/2026 // RR 2025/2026 YTD // Expected RR 2026 // … (61 total)',
     'Sutherland project variant — parallel order follow-up track with printer-specific OIT/RR views',
     'Data/Order-Intake/'),

    ('DPS', 'DPS_BP1 - AMSP Contribution Check (Final).xls', 'xls', 2,
     'BExRepositorySheet // Table',
     'SAP BW query — AMSP margin contribution by budget class/BU (BP1 plan), feeds margin % KPIs',
     'Data/Sales & Margin/'),

    ('DPS', 'DPS_BP1 - Sales details in all currencies.xls', 'xls', 2,
     'BExRepositorySheet // Table',
     'SAP BW query — detailed sales in all currencies (CO-PA), gross sales & net TO by currency',
     'Data/Sales & Margin/'),

    ('DPS', 'DPS_BP1_RECO Analysis Final.xls', 'xls', 2,
     'BExRepositorySheet // Table',
     'SAP BW query — RECO P&L waterfall (Revenue → Mfg Contribution → Gross Margin → SG&A → EBIT), IC elimination',
     'Data/Sales & Margin/'),

    ('DPS', 'DPS_BP2_Detailed Sales Inquiry AGFA NV_FY2025.xls', 'xls', 2,
     'BExRepositorySheet // Table',
     'SAP BW query — detailed sales inquiry under BP2 plan, FY2025 actuals & comparisons',
     'Data/Sales & Margin/'),

    # ── HE IT ──
    ('HE IT', 'OI HEC view pivot table (1).xlsx', 'xlsx', 15,
     'Pivot // S1 Regional // S1 Regional (2) // S1 Own IP Regional // S2 FA view // S3 FA view // S1 Geographical view // S1 FA view // S1 FA view (2) // S1 FA view (3) // OI per Business Type // S1 Bus Type Pivot // OI BUD26 per Bus Type // Pivot Bud Bus type // S1 Regional (2)',
     'Order Intake pivot — monthly/YTD/FY ACT/BUD/LY/FOR by BU, region, FA code, business type (13 snapshots from Access DB)',
     'Data/'),

    ('HE IT', '7.14 Order Book Overview pivot (BRM HQ views).xlsx', 'xlsx', 19,
     'Default pivot // Total OB evo // S1 per region // S1 per RevStr // S1 per RevStr (2) // S2 per region // S2 per RevStr // S4 per region // S4 per RevStr // IT per RevStr // IT per RevStr Grp // IT per RevStr (2) // IT per RevStr (3) // S1 OB EI // S1 OB EI per region // H1 per Region // S1 Regional (2) // S2 Regional (2) // BExRepositorySheet',
     'Order Book overview — timeline buckets (Planned CY, Overdue, Not Planned) by BU, region, revenue stream; SAP BW source',
     'Data/'),

    ('HE IT', 'Order Book detailed pivot.xlsm', 'xlsm', 3,
     'Sheet1 // Sheet2 // Home',
     'Order Book project-level detail — 39 columns per row, CRM IDs, planned receipt dates, VBA macros; full EUR (not kEUR)',
     'Data/'),

    ('HE IT', '20-TACO pivot 2025 Selectable x-rate.xlsm', 'xlsm', 4,
     'Pivot // Dashboard EUR // Source Report view // Report view',
     'TACO P&L pivot — actuals vs budget vs LY by FA line, selectable FX rate, 80+ company codes; key bottom-line is Line 85 EBIT',
     'Data/'),
]

# ── BUILD WORKBOOK ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

NCOLS = 7
WIDTHS = [4, 40, 20, 8, 58, 46, 24]

def make_sheet_header(ws, title, sub, colour):
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 20
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = Font(bold=True, color='FFFFFF', size=15, name='Calibri')
    c1.fill = fill(colour); c1.alignment = aln('center')
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = Font(color='FFFFFF', size=9, name='Calibri', italic=True)
    c2.fill = fill(colour); c2.alignment = aln('center')

def make_col_headers(ws, colour):
    hdrs = ['#', 'File Name', 'File Type', 'Sheets', 'Sheet / Page Names', 'Role / Purpose', 'Location in Folder']
    ws.row_dimensions[3].height = 18
    for i, (h, w) in enumerate(zip(hdrs, WIDTHS), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
        c.fill = fill(colour)
        c.alignment = aln('center', wrap=False)
        c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = 'A3:' + get_column_letter(NCOLS) + '3'

def build_dashboard_sheet(ws, title, sub, hdr_col, rows):
    make_sheet_header(ws, title, sub, hdr_col)
    make_col_headers(ws, hdr_col)
    r = 4
    for i, (_, fname, ftype, sheets, sheet_names, role, loc) in enumerate(rows, 1):
        ws.row_dimensions[r].height = 52
        fc, tc, label = TYPE_STYLE.get(ftype, (C['light'], '000000', ftype.upper()))
        wc(ws, r, 1, i, ha='center')
        wc(ws, r, 2, fname, bold=True)
        wc(ws, r, 3, label, bold=True, fc=fc, tc=tc, ha='center')
        count_val = sheets if sheets > 0 else '—'
        wc(ws, r, 4, count_val, bold=True, ha='center')
        wc(ws, r, 5, sheet_names, sz=8, italic=True, tc='444444')
        wc(ws, r, 6, role)
        wc(ws, r, 7, loc, sz=8, italic=True, fc=C['light'], tc='555555')
        r += 1

dashboards = [
    ('Digital Radiology', C['hdr_dr'],   '9 files  |  2 xlsx · 6 pbix · 1 docx  |  45 total sheets / pages'),
    ('Digital Hydrogen',  C['hdr_dh'],   '7 files  |  6 xlsx · 1 xls  |  132 total sheets'),
    ('DPS',               C['hdr_dps'],  '6 files  |  2 xlsx · 4 xls  |  131 total sheets'),
    ('HE IT',             C['hdr_heit'], '4 files  |  2 xlsx · 2 xlsm  |  41 total sheets'),
]
for dash_name, hdr_col, sub in dashboards:
    ws = wb.create_sheet(dash_name)
    rows = [d for d in DATA if d[0] == dash_name]
    build_dashboard_sheet(ws, dash_name + '  —  Source Files & Sheets', sub, hdr_col, rows)

# ── SUMMARY SHEET ──────────────────────────────────────────────────────────────
ws = wb.create_sheet('Summary', 0)
ws.row_dimensions[1].height = 34
ws.row_dimensions[2].height = 20
ws.merge_cells('A1:H1')
ws.merge_cells('A2:H2')
c1 = ws.cell(row=1, column=1, value='AGFA Analytics  —  Source Files & Sheets Register')
c1.font = Font(bold=True, color='FFFFFF', size=15, name='Calibri')
c1.fill = fill(C['hdr_sum']); c1.alignment = aln('center')
c2 = ws.cell(row=2, column=1, value='All source data files across Digital Radiology, Digital Hydrogen, DPS, HE IT  |  Generated 2026-03-27')
c2.font = Font(color='FFFFFF', size=9, name='Calibri', italic=True)
c2.fill = fill(C['hdr_sum']); c2.alignment = aln('center')

ws.row_dimensions[3].height = 10

sum_hdrs = ['Dashboard', 'Total Files', 'xlsx', 'xls', 'xlsm', 'pbix', 'docx', 'Total Sheets / Pages']
sum_widths = [24, 12, 8, 8, 8, 8, 8, 20]
ws.row_dimensions[4].height = 18
for i, (h, w) in enumerate(zip(sum_hdrs, sum_widths), 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
    c.fill = fill(C['hdr_sum']); c.alignment = aln('center', wrap=False); c.border = bdr()
    ws.column_dimensions[get_column_letter(i)].width = w

summary_rows = [
    ('Digital Radiology', 9,  2, 0, 0, 6, 1,  45,  C['sec_dr'],   C['hdr_dr']),
    ('Digital Hydrogen',  7,  6, 1, 0, 0, 0,  132, C['sec_dh'],   C['hdr_dh']),
    ('DPS',               6,  2, 4, 0, 0, 0,  131, C['sec_dps'],  C['hdr_dps']),
    ('HE IT',             4,  2, 0, 2, 0, 0,  41,  C['sec_heit'], C['hdr_heit']),
]

for i, (dash, files, xlsx, xls, xlsm, pbix, docx, total, sec, hdr) in enumerate(summary_rows, 5):
    ws.row_dimensions[i].height = 24
    wc(ws, i, 1, dash,  bold=True, fc=sec,          tc='1A2B4A')
    wc(ws, i, 2, files, bold=True, fc=C['sec_sum'], tc='1A2B4A', ha='center')
    wc(ws, i, 3, xlsx  if xlsx  else '', ha='center', fc=C['xlsx_c']  if xlsx  else C['light'], tc=C['t_xlsx']  if xlsx  else '999999')
    wc(ws, i, 4, xls   if xls   else '', ha='center', fc=C['xls_c']   if xls   else C['light'], tc=C['t_xls']   if xls   else '999999')
    wc(ws, i, 5, xlsm  if xlsm  else '', ha='center', fc=C['xlsm_c']  if xlsm  else C['light'], tc=C['t_xlsm']  if xlsm  else '999999')
    wc(ws, i, 6, pbix  if pbix  else '', ha='center', fc=C['pbix']    if pbix  else C['light'], tc=C['t_pbix']  if pbix  else '999999')
    wc(ws, i, 7, docx  if docx  else '', ha='center', fc=C['docx_c']  if docx  else C['light'], tc=C['t_docx']  if docx  else '999999')
    wc(ws, i, 8, total, bold=True, fc=sec, tc='1A2B4A', ha='center')

# Totals row
ws.row_dimensions[9].height = 24
for i, v in enumerate(['TOTAL', 26, 12, 5, 2, 6, 1, 349], 1):
    c = ws.cell(row=9, column=i, value=v)
    c.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    c.fill = fill(C['hdr_sum']); c.alignment = aln('center'); c.border = bdr()

# Legend
ws.row_dimensions[10].height = 10
ws.row_dimensions[11].height = 16
ws.merge_cells('A11:H11')
lc = ws.cell(row=11, column=1, value='  File Type Legend')
lc.font = Font(bold=True, color='1A2B4A', size=10, name='Calibri')
lc.fill = fill(C['sec_sum']); lc.alignment = aln(); lc.border = bdr()

legend = [
    ('Power BI Report (.pbix)',        C['pbix'],   C['t_pbix'],  'Interactive dashboard file — open with Power BI Desktop; parsed as ZIP to extract page names'),
    ('Excel Workbook (.xlsx)',         C['xlsx_c'], C['t_xlsx'],  'Standard Excel format — readable with openpyxl; contains pivot tables, raw data, and analysis sheets'),
    ('Excel Legacy / SAP BW (.xls)',   C['xls_c'],  C['t_xls'],  'SAP BW exported workbook — always contains BExRepositorySheet + Table layout; requires xlrd or win32com to read'),
    ('Excel Macro-Enabled (.xlsm)',    C['xlsm_c'], C['t_xlsm'], 'Excel with embedded VBA macros — openpyxl can read data but macros are not executed'),
    ('Word Document (.docx)',          C['docx_c'], C['t_docx'], 'Process/reference documentation — not a data source; provides business rules and field definitions'),
]
for i, (label, fc, tc, desc) in enumerate(legend, 12):
    ws.row_dimensions[i].height = 20
    wc(ws, i, 1, label, bold=True, fc=fc, tc=tc, ha='center')
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    wc(ws, i, 2, desc)
    for col in range(3, 9):
        ws.cell(row=i, column=col).border = bdr()

ws.freeze_panes = 'A5'

# ── Save ───────────────────────────────────────────────────────────────────────
out = r'c:\Users\vajra\OneDrive\Documents\Work\AGFA Analysis\AgfaAnalytics\Source_Files_Register.xlsx'
wb.save(out)
print('Saved:', out)
