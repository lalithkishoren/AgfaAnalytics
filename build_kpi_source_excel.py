import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    'hdr_dr':   '1565C0', 'sec_dr':   'E3F2FD',
    'hdr_dh':   '2E7D32', 'sec_dh':   'E8F5E9',
    'hdr_dps':  '6A1B9A', 'sec_dps':  'F3E5F5',
    'hdr_heit': 'E65100', 'sec_heit': 'FFF3E0',
    'hdr_sum':  '37474F', 'sec_sum':  'ECEFF1',
    'existing': 'E8F5E9', 't_exist':  '1B5E20',
    'new':      'FFF3E0', 't_new':    'E65100',
    'verified': 'E8F5E9', 't_ver':    '1B5E20',
    'derived':  'E3F2FD', 't_der':    '0D47A1',
    'estimated':'FFF9C4', 't_est':    'F57F17',
    'proxy':    'FFEBEE', 't_proxy':  'B71C1C',
    'white':    'FFFFFF', 'light':    'F5F7FA',
    'border':   'BDBDBD',
}

def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')
def aln(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style='thin', color=C['border'])
    return Border(left=s, right=s, top=s, bottom=s)
def wc(ws, r, c, val, bold=False, fc=None, tc='000000', sz=10, ha='left', wrap=True, italic=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = fnt(bold=bold, color=tc, size=sz, italic=italic)
    cell.alignment = aln(h=ha, wrap=wrap)
    if fc: cell.fill = fill(fc)
    cell.border = bdr()
    return cell

CONF_STYLE = {
    'Verified':  (C['verified'],  C['t_ver']),
    'Derived':   (C['derived'],   C['t_der']),
    'Estimated': (C['estimated'], C['t_est']),
    'Proxy':     (C['proxy'],     C['t_proxy']),
    'EDW Required': (C['proxy'],  C['t_proxy']),
    'Not Available': (C['proxy'], C['t_proxy']),
}

# ── DATA ───────────────────────────────────────────────────────────────────────
# (category, kpi_name, formula, source_in_client_report, confidence, origin, origin_detail)
# origin = 'Client Report' or 'Our Analysis'

DR_DATA = [
    # ── OIT ──
    ('Order Intake (OIT)', 'OIT YTD',
     'SUM(actualvalue_base) WHERE Won AND Year=2026',
     'Report 1 — OIT Margin & Product Mix (Power BI)',
     'Verified', 'Client Report', 'Already tracked in Power BI Report 1'),

    ('Order Intake (OIT)', 'Won Deals Count',
     'COUNT(opportunityid) WHERE Won AND Year=2026',
     'Report 1 — OIT Margin & Product Mix (Power BI)',
     'Verified', 'Client Report', 'Already tracked in Power BI Report 1'),

    ('Order Intake (OIT)', 'OIT vs Budget',
     '(OIT YTD / Budget YTD) − 1',
     'Report 1 — OIT Margin & Product Mix (Power BI)',
     'Estimated', 'Client Report', 'In Report 1; budget loaded as quarterly totals only'),

    ('Order Intake (OIT)', 'Avg Deal Size',
     'OIT YTD / Won Deals Count',
     'Not in any existing report',
     'Derived', 'Our Analysis', 'Simple average we derived — not in any Power BI report'),

    # ── Pipeline ──
    ('Pipeline & Funnel', 'Open Pipeline Total',
     'SUM(estimatedvalue_base) WHERE Open',
     'Report 1 & Report 2 (Power BI)',
     'Verified', 'Client Report', 'Tracked across multiple Power BI reports'),

    ('Pipeline & Funnel', 'RT CY / RT BT / RT PY',
     'Cumulative weekly OIT actual / budget / prior year',
     'Report 6 — OI & Funnel Health Cockpit (Power BI)',
     'Verified', 'Client Report', 'Running totals in Report 6'),

    ('Pipeline & Funnel', 'Weighted Amount',
     'SUM(agfa_weightedamountexcludingsma_base)',
     'Report 2 — Weekly FC Tracker (Power BI)',
     'Verified', 'Client Report', 'In Report 2 KAM Funnel view'),

    ('Pipeline & Funnel', 'Forecast Flag Snapshot',
     'Won / Incl&Secured / Included / InclwRisk / Upside / Excluded',
     'Report 2 — Weekly FC Tracker (Power BI)',
     'Verified', 'Client Report', 'Forecast flags from D365 forecast module'),

    ('Pipeline & Funnel', 'Pipeline Coverage',
     'Open Pipeline / (OIT YTD × 4) — target 2.5×',
     'Not in any existing report',
     'Derived', 'Our Analysis', 'Playbook mentions 2.5× target but no report calculates this ratio'),

    ('Pipeline & Funnel', '2× Upside Rule',
     'Upside ÷ Included with Risk ≥ 2.0 per region',
     'Not in any existing report',
     'Estimated', 'Our Analysis', 'Rule exists in Sales Playbook but never visualised in any report'),

    # ── Deal Scoring ──
    ('Deal Scoring', 'DS% (Deal Sign)',
     'agfa_dsdealsigncodename — CRM field',
     'Report 1 & Report 2 (Power BI)',
     'Verified', 'Client Report', 'CRM score surfaced in Power BI reports'),

    ('Deal Scoring', 'DH% (Deal Happen)',
     'agfa_dhdealhappencodename — CRM field',
     'Report 1 & Report 2 (Power BI)',
     'Verified', 'Client Report', 'CRM score surfaced in Power BI reports'),

    ('Deal Scoring', 'Feasibility %',
     'agfa_feasibilitycode — CRM field (0–90 scale)',
     'Report 1 & Report 2 (Power BI)',
     'Verified', 'Client Report', 'CRM score surfaced in Power BI reports'),

    ('Deal Scoring', 'Forecast Flag',
     'msdyn_forecastcategoryname — D365 forecast module',
     'Report 2 — Weekly FC Tracker (Power BI)',
     'Verified', 'Client Report', 'Core forecast flag shown in all FC views'),

    # ── Margin ──
    ('Margin', 'CRM Margin % (Total)',
     'agfa_margincostpercentagetotal — KAM-entered',
     'Report 1 — OIT Margin & Product Mix (Power BI)',
     'Estimated', 'Client Report', 'KAM estimate shown in Report 1; not SAP actual'),

    ('Margin', 'HW / Impl / License / Service Margin %',
     'agfa_margincostpercentage[hardware/implementation/licenses/services]',
     'Report 1 — OIT Margin & Product Mix (Power BI)',
     'Estimated', 'Client Report', 'Component margins from Sofon, shown in Report 1'),

    ('Margin', 'Actual SAP Margin',
     '(Net Turnover EUR − Calculated Cost APX) / Net Turnover EUR',
     'Report 3 — Partner Dashboard (Power BI)',
     'Verified', 'Client Report', 'SAP posted cost in FeedFile; shown in Report 3'),

    ('Margin', 'Standard Margin (Sofon)',
     '(Net Turnover − Sofon Cost+) / Net Turnover',
     'Report 4 — Price Margin Modalities (DirectQuery)',
     'Proxy', 'Client Report', 'Exists as DirectQuery in Report 4 but not extractable/connected in React dashboard'),

    # ── Revenue / Order Book ──
    ('Revenue / Order Book', 'Won with SAP Order ID',
     'COUNT WHERE agfa_saporderid NOT NULL AND Won',
     'Not in any existing report',
     'Derived', 'Our Analysis', 'We identified 1,716 Won deals have SAP Order ID — bridge to SAP never built'),

    ('Revenue / Order Book', 'Planned Reco Count',
     'COUNT WHERE planned_reco_date NOT NULL AND Open',
     'Not in any existing report',
     'Estimated', 'Our Analysis', 'Field exists in CRM (2,451 deals have it) but never surfaced in any report'),

    ('Revenue / Order Book', 'Order Book (Backlog)',
     'SUM(Won order value) WHERE not yet invoiced in SAP',
     'Not in any existing report — EDW required',
     'EDW Required', 'Our Analysis', 'Not available in any current report; requires agfa_saporderid → SAP order status join'),

    ('Revenue / Order Book', 'Actual Revenue Reco',
     'SUM(Net Turnover EUR) by Invoice Posting Date',
     'Not in any existing report — EDW required',
     'EDW Required', 'Our Analysis', 'FeedFile has posting dates but cannot be joined to CRM deal-level planned reco dates'),
]

DH_DATA = [
    # ── Revenue & Orders ──
    ('Revenue & Orders', 'YTD Revenue',
     'SUM(amount) WHERE currency=EUR, status=2, year=2026',
     'Sales zirfon GHS.xlsx — annual tabs',
     'Verified', 'Client Report', 'Raw data existed in GHS Excel; we structured it into orders.json'),

    ('Revenue & Orders', 'Full Year Forecast',
     'Actuals + Committed + Uncommitted + Unidentified',
     'ACTFY2025_Forecasting file.xlsx & Sales Forecast Feb2026.xlsx',
     'Verified', 'Client Report', 'Forecast composition existed in forecasting workbooks'),

    ('Revenue & Orders', 'vs Budget %',
     '(FY Forecast − €17.3M) / €17.3M × 100',
     'Sales Forecast Feb2026.xlsx — FOR vs BUD & LY sheet',
     'Verified', 'Client Report', 'Budget comparison existed in forecasting file'),

    ('Revenue & Orders', 'vs Prior Year %',
     '(FY Forecast − LY) / LY × 100',
     'FY 2025.xls — SAP BW workbook',
     'Verified', 'Client Report', 'LY actuals in FY 2025.xls; comparison existed'),

    ('Revenue & Orders', 'Revenue by Product',
     'EUR split by UTP500 / UTP220 / UTP500+',
     'Sales zirfon GHS.xlsx — annual tabs',
     'Verified', 'Client Report', 'Product breakdown existed in GHS annual sheets'),

    ('Revenue & Orders', 'Customer Concentration (Top 5/10)',
     'Top 5 / Top 10 as % of total EUR revenue',
     'Not in any existing report',
     'Derived', 'Our Analysis', 'We ranked and computed concentration — not done in any source file'),

    ('Revenue & Orders', 'Open Orders Count & Value',
     'COUNT / SUM WHERE status=1',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'Status field exists in GHS; we filtered and surfaced this view'),

    ('Revenue & Orders', 'Average Order Value',
     'Total EUR / completed order count',
     'Not in any existing report',
     'Derived', 'Our Analysis', 'We computed this — not in any source Excel'),

    # ── Pipeline & Conversion ──
    ('Pipeline & Conversion', 'Quote Conversion Rate',
     'isOrdered=true / Total Quotations (overall: 18.0%)',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We joined Quotation.xlsx → GHS to compute conversion'),

    ('Pipeline & Conversion', 'Pipeline Value & Count',
     'SUM/COUNT WHERE isOrdered=false',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We derived the pipeline from non-ordered quotes'),

    ('Pipeline & Conversion', 'Conversion Rate by Year',
     '2023: 8.8% → 2024: 24.1% → 2025: 24.8%',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We segmented conversion per year — not done in source files'),

    ('Pipeline & Conversion', 'Conversion Rate by Product',
     'UTP220: 22.9% · UTP500+: 16.5% · UTP500: 16.2%',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We segmented conversion per product type'),

    ('Pipeline & Conversion', 'Avg Quote Size (Ordered vs Non-Ordered)',
     'Ordered avg €20,856 vs Non-ordered avg €197,046',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We compared two populations — key insight not in any source file'),

    # ── Margin ──
    ('Margin & Profitability', 'Gross Margin %',
     '(Net Turnover − Std Cost) / Net Turnover × 100',
     'FY 2025.xls — Month Margin / Full Pivot Margin sheets',
     'Verified', 'Client Report', 'Margin existed in SAP BW workbook; we extracted and structured it'),

    ('Margin & Profitability', 'Standard Cost Prices',
     'UTP500: €115.86/m² · UTP220: €95.47/m² · UTP500+: €115.86/m²',
     'FY 2025.xls — Mapping Standard Costprices sheet',
     'Verified', 'Client Report', 'Standard costs existed in dedicated mapping sheet'),

    ('Margin & Profitability', 'MSP 2026',
     'UTP220: €111.07/m² · UTP500: €102.10/m²',
     'Sales Forecast Feb2026.xlsx — Revenue overview sheet',
     'Verified', 'Client Report', 'MSP values existed in Revenue overview tab'),

    ('Margin & Profitability', 'Average Selling Price (EUR/m²)',
     'EUR Revenue / Total m² sold, by product and year',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We computed EUR/m² by dividing revenue by volume — not in source files'),

    ('Margin & Profitability', 'Gross Margin (Absolute)',
     'Net Turnover − Standard Cost Total (EUR)',
     'FY 2025.xls — Month Margin sheet',
     'Verified', 'Client Report', 'Existed in SAP BW margin pivot'),

    # ── Forecast & Plans ──
    ('Forecast & Plans', 'Forecast Composition',
     'Actuals 27% + Committed 35% + Uncommitted 13% + Unidentified 25%',
     'ACTFY2025_Forecasting file.xlsx — FOR Summary sheet',
     'Estimated', 'Client Report', 'Forecast type breakdown existed in forecasting workbook'),

    ('Forecast & Plans', 'Budget Gap',
     'FY2026 Forecast €6.2M − Budget €17.3M = −€11.1M (−64.3%)',
     'Sales Forecast Feb2026.xlsx — FOR vs BUD & LY sheet',
     'Verified', 'Client Report', 'Budget gap comparison existed in forecasting file'),

    ('Forecast & Plans', 'Forecast Revision History',
     'BUD €17.3M → RFC2 €31.2M → Current €6.2M',
     'ACTFY2025 & Feb2026 forecasting files',
     'Verified', 'Client Report', 'Revision history existed across forecasting workbooks'),

    ('Forecast & Plans', 'Long-Term Plans (2027–2029)',
     '2027: €38.6M · 2028: €51.4M · 2029: €69.9M',
     'Sales Forecast Feb2026.xlsx — Revenue overview sheet',
     'Estimated', 'Client Report', 'LTP existed in Revenue overview tab'),

    ('Forecast & Plans', 'Scenario Analysis (Bear/Base/Bull)',
     'Base / Bear ×0.7 / Bull ×1.2',
     'Not in any existing report',
     'Estimated', 'Our Analysis', 'We built the scenario model with sliders — not in any source file'),

    ('Forecast & Plans', 'TK Nucera NEOM Batches',
     'PO 32017233 — Batches #6–#25, ~€949K each',
     'ACTFY2025_Forecasting file.xlsx — TK Nucera sheet',
     'Verified', 'Client Report', 'TK Nucera batch data existed in dedicated sheet'),

    # ── Customer 360 ──
    ('Customer 360°', 'Lifetime Revenue per Customer',
     'SUM(EUR) per customer WHERE status=2, all years',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We aggregated per customer across all years — not in any source file'),

    ('Customer 360°', 'Quotation Conversion per Customer',
     'Converted / Total quotes per customer',
     'Not in any existing report',
     'Verified', 'Our Analysis', 'We computed per-customer conversion — not in any source file'),

    ('Customer 360°', 'Product Mix per Customer',
     'Revenue by product type for selected customer',
     'Not in any existing report',
     'Derived', 'Our Analysis', 'We built customer-level product breakdown — not in any source file'),
]

DPS_DATA = [
    ('Revenue & Margin', 'Net Revenue (3rd Party)',
     'Sum of Net Sales to third-party customers',
     'DPS_BP1 - AMSP Contribution Check (Final).xls',
     'Verified', 'Client Report', 'Core metric from SAP BW AMSP report'),

    ('Revenue & Margin', 'AMSP Margin %',
     'AMSP Contribution / Net Sales 3rdP × 100',
     'DPS_BP1 - AMSP Contribution Check (Final).xls',
     'Verified', 'Client Report', 'Core margin metric from SAP BW AMSP report'),

    ('Revenue & Margin', 'CO-PA Gross Margin %',
     'Gross Margin / Net TO × 100',
     'DPS_BP1 - Sales details in all currencies.xls (CO-PA)',
     'Verified', 'Client Report', 'From SAP CO-PA GMPCOPA_1 query'),

    ('Revenue & Margin', 'EBIT',
     'Adjusted EBIT + Non-recurring items',
     'DPS_BP1_RECO Analysis Final.xls',
     'Verified', 'Client Report', 'Bottom-line EBIT from SAP BW RECO waterfall'),

    ('Revenue & Margin', 'Budget Variance %',
     '(Actual − Budget) / Budget × 100',
     'DPS_BP1_RECO Analysis Final.xls — RECO vs BP1',
     'Derived', 'Client Report', 'Variance existed in RECO report vs BP1 budget column'),

    ('Order Intake & Pipeline', 'OIT Units',
     'Count of order intake records by period',
     'DPS_Customer order & revenue follow-up 2026.xlsx',
     'Verified', 'Client Report', 'Tracked manually in order follow-up Excel per month'),

    ('Order Intake & Pipeline', 'RR Units',
     'Count of invoiced/recognised orders by period',
     'DPS_Customer order & revenue follow-up 2026.xlsx',
     'Verified', 'Client Report', 'RR tracked per month in order follow-up Excel'),

    ('Order Intake & Pipeline', 'Delayed Orders',
     'Count of orders with revised RR > original RR',
     'DPS_Customer order & revenue follow-up 2026.xlsx — Delayed tab',
     'Verified', 'Client Report', 'Dedicated Delayed tracker tab in the Excel'),

    ('Order Intake & Pipeline', 'Backlog Units',
     'OIT cumulative − RR cumulative',
     'DPS_Customer order & revenue follow-up 2026.xlsx',
     'Derived', 'Client Report', 'Derivable from OIT and RR columns in order follow-up'),

    ('Order Intake & Pipeline', 'OIT EUR Value',
     'Sum of order values at intake date',
     'NOT AVAILABLE — order follow-up tracks units only',
     'Not Available', 'Our Analysis', 'We flagged this gap — no EUR value in order follow-up; SAP link needed'),

    ('Order Intake & Pipeline', 'Pipeline EUR Value',
     'Sum of forward pipeline order values',
     'NOT AVAILABLE — no CRM or pipeline EUR for DPS',
     'Not Available', 'Our Analysis', 'We flagged this gap — no pipeline EUR data exists for DPS'),

    ('Order Intake & Pipeline', 'HW vs Consumable Split (EUR)',
     'Revenue by Financial Application dimension',
     'NOT AVAILABLE — FA dimension missing from BW extract',
     'Not Available', 'Our Analysis', 'We flagged this gap — BW query does not include FA dimension filter'),
]

HEIT_DATA = [
    ('OI — Order Intake', 'MONTH ACT',
     'Monthly OI actual — new orders booked in the month',
     'OI HEC view pivot table (1).xlsx — Pivot sheet (Key Figure = MONTH ACT)',
     'Verified', 'Client Report', 'Core key figure from OI Access database'),

    ('OI — Order Intake', 'MONTH BUD',
     'Monthly OI budget — annual plan phased monthly',
     'OI HEC view pivot table (1).xlsx — Pivot sheet (Key Figure = MONTH BUD)',
     'Verified', 'Client Report', 'Core key figure from OI Access database'),

    ('OI — Order Intake', 'MONTH LY',
     'Monthly OI prior year — same calendar month, prior year',
     'OI HEC view pivot table (1).xlsx — Pivot sheet (Key Figure = MONTH LY)',
     'Verified', 'Client Report', 'Core key figure from OI Access database'),

    ('OI — Order Intake', 'MONTH FOR',
     'Monthly OI forecast — actuals YTD + forecast balance',
     'OI HEC view pivot table (1).xlsx — Pivot sheet (Key Figure = MONTH FOR)',
     'Verified', 'Our Analysis', 'Data EXISTS in pivot (13 snapshots) but NOT surfaced in any UI chart — we flagged as P2 quick win'),

    ('OI — Order Intake', 'YTD ACT',
     'Year-to-date cumulative OI actuals Jan → reporting month',
     'OI HEC view pivot table (1).xlsx — oi_ytd.json (Key Figure = YTD ACT)',
     'Verified', 'Client Report', 'Core YTD key figure from OI Access database'),

    ('OI — Order Intake', 'YTD BUD',
     'Year-to-date OI budget cumulated to same period',
     'OI HEC view pivot table (1).xlsx — oi_ytd.json (Key Figure = YTD BUD)',
     'Verified', 'Client Report', 'Core YTD key figure from OI Access database'),

    ('OI — Order Intake', 'YTD LY',
     'Year-to-date OI prior year cumulated to same period',
     'OI HEC view pivot table (1).xlsx — oi_ytd.json (Key Figure = YTD LY)',
     'Verified', 'Client Report', 'Core YTD key figure from OI Access database'),

    ('OI — Order Intake', 'FY (A+F)',
     'Full-year Actual + Forecast — actuals YTD + remaining forecast',
     'OI HEC view pivot table (1).xlsx (Key Figure = FY (A+F))',
     'Verified', 'Client Report', 'Full-year outlook key figure from OI Access database'),

    ('OI — Order Intake', 'FY BUD',
     'Full-year OI budget (static annual plan)',
     'OI HEC view pivot table (1).xlsx (Key Figure = FY BUD)',
     'Verified', 'Client Report', 'Full-year budget key figure from OI Access database'),

    ('OB — Order Book', 'Planned Current Year',
     'Backlog planned for recognition in current FY',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx — Total OB evo sheet',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database — Low risk'),

    ('OB — Order Book', 'Planned Next Years',
     'Backlog planned beyond current FY (multi-year)',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database — Medium risk'),

    ('OB — Order Book', 'Rev Overdue ≤ 6 months',
     'Revenue planned in past period, not yet recognised ≤6m',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database — High risk'),

    ('OB — Order Book', 'Rev Overdue > 6 months',
     'Revenue overdue >6 months — critical delivery/contract risk',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database — Critical'),

    ('OB — Order Book', 'Not Planned Sales Order',
     'Sales orders received but not yet scheduled for recognition',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database — Medium risk'),

    ('OB — Order Book', 'Not Planned Opportunity ≤ 3 months',
     'Near-term unplanned opportunity-stage orders',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database'),

    ('OB — Order Book', 'Not Planned Opportunity 4–6 months',
     'Mid-range unplanned opportunities',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database'),

    ('OB — Order Book', 'Not Planned Opportunity 7–24 months',
     'Longer-horizon unplanned opportunities',
     '7.14 Order Book Overview pivot (BRM HQ views).xlsx',
     'Verified', 'Client Report', 'OB bucket from Order Book Access database'),

    ('TACO — P&L', 'Line 02 — Net Sales Hardware/Equipment',
     'Revenue from hardware equipment sales (net of returns/discounts)',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 07 — Net Sales Own Licenses',
     'Revenue from proprietary software licenses',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 09 — Net Sales subs. Own IP',
     'Subscription revenue for own IP software',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 11 — Net Sales 3rd Party Licenses',
     'Revenue from reselling third-party software licenses',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 26 — Net Sales Total',
     'Total revenue across all categories (sum of lines 02–25)',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 55 — TACO Margin',
     'Gross profit after direct costs. TACO Margin % = Line 55 / Line 26',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 63 — Product Contribution',
     'TACO Margin minus product-specific overheads',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key P&L line from TACO Access database'),

    ('TACO — P&L', 'Line 85 — TACO Contribution (EBIT)',
     'Operating result after all allocated expenses — key bottom line',
     '20-TACO pivot 2025 Selectable x-rate.xlsm — Pivot sheet',
     'Verified', 'Client Report', 'Key bottom-line P&L from TACO Access database'),
]

# ── BUILD ──────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

NCOLS = 8
WIDTHS = [4, 22, 28, 42, 38, 14, 16, 44]
HEADERS = ['#', 'Category', 'KPI Name', 'Formula / Definition',
           'Source in Client Report / File', 'Confidence', 'Origin', 'Notes']

def make_header(ws, title, sub, colour):
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 20
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = Font(bold=True, color='FFFFFF', size=15, name='Calibri')
    c1.fill = fill(colour); c1.alignment = aln('center')
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = Font(color='FFFFFF', size=9, name='Calibri', italic=True)
    c2.fill = fill(colour); c2.alignment = aln('center')

def make_col_headers(ws, colour):
    ws.row_dimensions[3].height = 18
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
        c.fill = fill(colour); c.alignment = aln('center', wrap=False); c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = 'A3:' + get_column_letter(NCOLS) + '3'

def section_row(ws, row, label, sec_colour):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    cell = ws.cell(row=row, column=1, value=f'  {label}')
    cell.font = Font(bold=True, color='1A2B4A', size=10, name='Calibri')
    cell.fill = fill(sec_colour); cell.alignment = aln('left', wrap=False); cell.border = bdr()

def build_sheet(ws, title, sub, hdr_col, sec_col, data):
    make_header(ws, title, sub, hdr_col)
    make_col_headers(ws, hdr_col)
    row = 4
    num = 0
    last_cat = None
    for (cat, kpi, formula, source, conf, origin, notes) in data:
        if cat != last_cat:
            section_row(ws, row, cat, sec_col)
            row += 1
            last_cat = cat
        num += 1
        ws.row_dimensions[row].height = 42

        # Origin badge
        if origin == 'Client Report':
            ofc, otc = C['existing'], C['t_exist']
        else:
            ofc, otc = C['new'], C['t_new']

        # Confidence badge
        conf_key = conf.split()[0] if conf else ''
        cfc, ctc = CONF_STYLE.get(conf_key, (C['light'], '000000'))
        if conf in ('Not Available', 'EDW Required'):
            cfc, ctc = C['proxy'], C['t_proxy']

        wc(ws, row, 1, num, ha='center')
        wc(ws, row, 2, cat)
        wc(ws, row, 3, kpi, bold=True)
        wc(ws, row, 4, formula, sz=9)
        wc(ws, row, 5, source, sz=9, italic=True, tc='444444')
        wc(ws, row, 6, conf, bold=True, fc=cfc, tc=ctc, ha='center')
        wc(ws, row, 7, origin, bold=True, fc=ofc, tc=otc, ha='center')
        wc(ws, row, 8, notes, sz=9, italic=True, tc='555555')
        row += 1
    return row

# Build dashboard sheets
sheets = [
    ('Digital Radiology', C['hdr_dr'], C['sec_dr'],
     '22 KPIs defined  |  14 from client reports  |  8 from our analysis',
     DR_DATA),
    ('Digital Hydrogen', C['hdr_dh'], C['sec_dh'],
     '30 KPIs defined  |  12 from client reports  |  18 from our analysis',
     DH_DATA),
    ('DPS', C['hdr_dps'], C['sec_dps'],
     '12 KPIs defined  |  9 from client reports  |  3 from our analysis',
     DPS_DATA),
    ('HE IT', C['hdr_heit'], C['sec_heit'],
     '26 KPIs defined  |  25 from client reports  |  1 from our analysis',
     HEIT_DATA),
]

for name, hdr, sec, sub, data in sheets:
    ws = wb.create_sheet(name)
    build_sheet(ws, name + '  —  KPI Definitions: Client Report vs Our Analysis', sub, hdr, sec, data)

# ── SUMMARY SHEET ──────────────────────────────────────────────────────────────
ws = wb.create_sheet('Summary', 0)
ws.row_dimensions[1].height = 34
ws.row_dimensions[2].height = 20
ws.merge_cells('A1:G1'); ws.merge_cells('A2:G2')
c1 = ws.cell(row=1, column=1, value='KPI Definitions — Client Report vs Our Analysis')
c1.font = Font(bold=True, color='FFFFFF', size=15, name='Calibri')
c1.fill = fill(C['hdr_sum']); c1.alignment = aln('center')
c2 = ws.cell(row=2, column=1, value='Which KPIs were already in existing reports vs which we identified through data analysis  |  Generated 2026-03-27')
c2.font = Font(color='FFFFFF', size=9, name='Calibri', italic=True)
c2.fill = fill(C['hdr_sum']); c2.alignment = aln('center')

ws.row_dimensions[3].height = 10
ws.row_dimensions[4].height = 18

sum_hdrs = ['Dashboard', 'Total KPIs', 'From Client Reports', 'From Our Analysis', '% Already Tracked', '% We Added', '']
sum_widths = [24, 12, 20, 18, 18, 14, 4]
for i, (h, w) in enumerate(zip(sum_hdrs, sum_widths), 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=9, name='Calibri')
    c.fill = fill(C['hdr_sum']); c.alignment = aln('center', wrap=False); c.border = bdr()
    ws.column_dimensions[get_column_letter(i)].width = w

summary_rows = [
    ('Digital Radiology', 22, 14, 8,  C['sec_dr'],   C['hdr_dr']),
    ('Digital Hydrogen',  30, 12, 18, C['sec_dh'],   C['hdr_dh']),
    ('DPS',               12,  9,  3, C['sec_dps'],  C['hdr_dps']),
    ('HE IT',             26, 25,  1, C['sec_heit'], C['hdr_heit']),
]

for i, (dash, total, client, ours, sec, hdr) in enumerate(summary_rows, 5):
    ws.row_dimensions[i].height = 24
    pct_client = f'{round(client/total*100)}%'
    pct_ours   = f'{round(ours/total*100)}%'
    wc(ws, i, 1, dash,       bold=True, fc=sec,         tc='1A2B4A')
    wc(ws, i, 2, total,      bold=True, fc=C['sec_sum'], tc='1A2B4A', ha='center')
    wc(ws, i, 3, client,     bold=True, fc=C['existing'],tc=C['t_exist'], ha='center')
    wc(ws, i, 4, ours,       bold=True, fc=C['new'],    tc=C['t_new'], ha='center')
    wc(ws, i, 5, pct_client, fc=C['existing'], tc=C['t_exist'], ha='center')
    wc(ws, i, 6, pct_ours,   fc=C['new'],      tc=C['t_new'],   ha='center')
    ws.cell(row=i, column=7).fill = fill(C['white'])

# Totals
ws.row_dimensions[9].height = 24
for i, v in enumerate(['TOTAL', 90, 60, 30, '67%', '33%', ''], 1):
    c = ws.cell(row=9, column=i, value=v)
    c.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    c.fill = fill(C['hdr_sum']); c.alignment = aln('center'); c.border = bdr()

# Legend
ws.row_dimensions[10].height = 10
ws.row_dimensions[11].height = 16
ws.merge_cells('A11:G11')
lc = ws.cell(row=11, column=1, value='  Origin Legend')
lc.font = Font(bold=True, color='1A2B4A', size=10, name='Calibri')
lc.fill = fill(C['sec_sum']); lc.alignment = aln(); lc.border = bdr()

legend = [
    ('Client Report', C['existing'], C['t_exist'],
     'KPI was already tracked in the client\'s existing Excel files or Power BI reports — we documented and structured it'),
    ('Our Analysis',  C['new'],      C['t_new'],
     'KPI was identified, derived, or flagged through our data analysis work — not present in any source file or report'),
]
for i, (label, fc, tc, desc) in enumerate(legend, 12):
    ws.row_dimensions[i].height = 20
    wc(ws, i, 1, label, bold=True, fc=fc, tc=tc, ha='center')
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
    wc(ws, i, 2, desc)
    for col in range(3, 8):
        ws.cell(row=i, column=col).border = bdr()

ws.freeze_panes = 'A5'

out = r'c:\Users\vajra\OneDrive\Documents\Work\AGFA Analysis\AgfaAnalytics\KPI_Source_Analysis.xlsx'
wb.save(out)
print('Saved:', out)
