import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
C = {
    "hdr_dr":     "1565C0",
    "hdr_dh":     "2E7D32",
    "hdr_dps":    "6A1B9A",
    "hdr_heit":   "E65100",
    "hdr_sum":    "37474F",

    "sec_dr":     "E3F2FD",
    "sec_dh":     "E8F5E9",
    "sec_dps":    "F3E5F5",
    "sec_heit":   "FFF3E0",
    "sec_sum":    "ECEFF1",

    "verified":   "E8F5E9",
    "derived":    "E3F2FD",
    "estimated":  "FFF9C4",
    "proxy":      "FFEBEE",
    "proposed":   "FFF3E0",
    "edw":        "FFEBEE",
    "ready":      "E8F5E9",
    "new_type":   "FFF3E0",
    "exist_type": "E8F5E9",

    "t_verified": "1B5E20",
    "t_derived":  "0D47A1",
    "t_estimated":"F57F17",
    "t_proxy":    "B71C1C",
    "t_proposed": "E65100",
    "t_edw":      "B71C1C",
    "t_ready":    "1B5E20",
    "t_new":      "E65100",
    "t_exist":    "1B5E20",

    "white":      "FFFFFF",
    "light_grey": "F5F7FA",
    "border":     "BDBDBD",
}

def fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def write_cell(ws, row, col, value, bold=False, fcolor=None, tcolor="000000",
               size=10, h_align="left", wrap=True, italic=False, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, color=tcolor, size=size, italic=italic)
    cell.alignment = align(h=h_align, wrap=wrap)
    if fcolor:
        cell.fill = fill(fcolor)
    if border:
        cell.border = thin_border()
    return cell

def sheet_header(ws, title, subtitle, colour, col_count):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = Font(bold=True, color=C["white"], size=16, name="Calibri")
    c1.fill = fill(colour)
    c1.alignment = align(h="center", v="center")
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(bold=False, color=C["white"], size=10, name="Calibri", italic=True)
    c2.fill = fill(colour)
    c2.alignment = align(h="center", v="center")

def col_header_row(ws, row, headers, colours, widths=None):
    ws.row_dimensions[row].height = 18
    for i, (h, c) in enumerate(zip(headers, colours), start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, color=C["white"], size=9, name="Calibri")
        cell.fill = fill(c)
        cell.alignment = align(h="center", v="center", wrap=False)
        cell.border = thin_border()
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def section_row(ws, row, label, colour, col_count):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color="1A2B4A", size=10, name="Calibri")
    cell.fill = fill(colour)
    cell.alignment = align(h="left", v="center", wrap=False)
    cell.border = thin_border()

def divider_row(ws, row, label, colour, col_count):
    """Bold full-width divider to separate Existing vs New blocks."""
    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color=C["white"], size=11, name="Calibri")
    cell.fill = fill(colour)
    cell.alignment = align(h="center", v="center", wrap=False)
    cell.border = thin_border()

# Confidence / status auto-colour (applied to col 8 — "Confidence / Status")
CONF_STYLE = {
    "✅ Verified":  (C["verified"],  C["t_verified"]),
    "🔵 Derived":   (C["derived"],   C["t_derived"]),
    "⚠ Estimated":  (C["estimated"], C["t_estimated"]),
    "🔴 Proxy":     (C["proxy"],     C["t_proxy"]),
    "🔴 EDW":       (C["edw"],       C["t_edw"]),
    "✅ Ready":     (C["ready"],     C["t_ready"]),
    "⚠ Partial":   (C["estimated"], C["t_estimated"]),
}

# Priority auto-colour (col 9)
PRI_STYLE = {
    "P1": (C["proxy"],     C["t_proxy"]),
    "P2": (C["verified"],  C["t_verified"]),
    "P3": (C["estimated"], C["t_estimated"]),
    "P4": (C["estimated"], C["t_estimated"]),
    "P5": (C["derived"],   C["t_derived"]),
    "P6": (C["derived"],   C["t_derived"]),
    "P7": (C["proxy"],     C["t_proxy"]),
}

# ═══════════════════════════════════════════════════════════════════════════════
# DATA — Existing KPIs  (section, kpi, formula, source, unit, confidence)
# ═══════════════════════════════════════════════════════════════════════════════

DR_EXISTING = [
    ("Order Intake (OIT)", "OIT YTD 2026", "SUM(actualvalue_base) WHERE Won AND Year=2026", "msd data.csv", "EUR", "✅ Verified"),
    ("Order Intake (OIT)", "Won Deals Count", "COUNT(opportunityid) WHERE Won AND Year=2026", "msd data.csv", "Count", "✅ Verified"),
    ("Order Intake (OIT)", "Avg Deal Size", "OIT YTD / Won Deals Count", "Derived", "EUR", "🔵 Derived"),
    ("Order Intake (OIT)", "OIT vs Budget %", "(OIT YTD / Budget YTD) − 1", "msd data + Budget Quarter.csv", "%", "⚠ Estimated"),
    ("Order Intake (OIT)", "OIT FY 2025 / FY 2024", "Historical actuals", "msd data.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "Open Pipeline Total", "SUM(estimatedvalue_base) WHERE Open", "opportunity.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "Open Pipeline Count", "COUNT(opportunityid) WHERE Open", "opportunity.csv", "Count", "✅ Verified"),
    ("Pipeline & Funnel", "Pipeline Coverage", "Open Pipeline / (OIT YTD × 4)", "Derived", "Ratio", "🔵 Derived"),
    ("Pipeline & Funnel", "RT CY (Running Total CY)", "Cumulative weekly OIT actuals", "T funnel health.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "RT BT (Running Total Budget)", "Cumulative weekly OIT budget", "T funnel health.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "RT PY (Running Total PY)", "Cumulative weekly OIT prior year", "T funnel health.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "Weighted Amount", "SUM(agfa_weightedamountexcludingsma_base)", "DataWeek.csv / opportunity.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "Forecast Flag Snapshot", "Won/Incl&Secured/Included/InclwRisk/Upside/Excluded", "DataWeek.csv", "EUR", "✅ Verified"),
    ("Pipeline & Funnel", "Funnel Evolution (weekly)", "Funnel by flag per week × region", "T funnel evolution tracker.csv", "EUR", "✅ Verified"),
    ("Deal Scoring", "DS% (Deal Sign)", "agfa_dsdealsigncodename", "opportunity.csv", "Score", "✅ Verified"),
    ("Deal Scoring", "DH% (Deal Happen)", "agfa_dhdealhappencodename", "opportunity.csv", "Score", "✅ Verified"),
    ("Deal Scoring", "Feasibility %", "agfa_feasibilitycode (0–90 integer scale)", "opportunity.csv", "Score", "✅ Verified"),
    ("Margin", "CRM Margin % (Total)", "agfa_margincostpercentagetotal — KAM-entered", "msd data.csv", "%", "⚠ Estimated"),
    ("Margin", "Hardware Margin %", "agfa_margincostpercentagehardware", "msd data.csv", "%", "⚠ Estimated"),
    ("Margin", "Implementation Margin %", "agfa_margincostpercentageimplementation", "msd data.csv", "%", "⚠ Estimated"),
    ("Margin", "License Margin %", "agfa_margincostpercentageinternallicenses", "msd data.csv", "%", "⚠ Estimated"),
    ("Margin", "Service Margin %", "agfa_margincostpercentageservicecontracts", "msd data.csv", "%", "⚠ Estimated"),
    ("Margin", "SAP Actual Margin (Goods)", "(Net Turnover EUR − Calculated Cost APX) / Net TO", "FeedFile.csv (Report 2)", "%", "✅ Verified"),
    ("Margin", "Standard Margin (Sofon)", "(Net Turnover − Sofon Cost+) / Net Turnover", "Q price realisation (DirectQuery)", "%", "🔴 Proxy"),
    ("Revenue / Actuals", "Partner Revenue vs Target", "Dealer actual EUR vs budget by quarter", "FeedFile + DealerList_TargetSetting", "EUR", "✅ Verified"),
    ("Revenue / Actuals", "Revenue by Year & Type", "SAP actuals: Goods / Implementation / Support", "FeedFile.csv", "EUR", "✅ Verified"),
    ("Revenue / Actuals", "Price Realization Waterfall", "List Price → Discount → Net TO → Sofon Cost+ → Gross Margin", "Q price realisation (Report 3)", "EUR", "✅ Verified"),
    ("Revenue / Actuals", "ENP (Effective Net Price)", "Per-unit pricing from AP2", "BP2 query", "EUR/unit", "✅ Verified"),
    ("Revenue / Actuals", "SAP Order Count (Won)", "COUNT WHERE agfa_saporderid IS NOT NULL AND Won", "msd data.csv", "Count", "🔵 Derived"),
    ("Revenue / Actuals", "Planned Reco Count", "COUNT WHERE planned_reco_date IS NOT NULL AND Open", "msd data.csv", "Count", "⚠ Estimated"),
]

DR_PROPOSED = [
    # (section, kpi, description, data_available, priority, status)
    ("Revenue Recognition", "Planned Reco by Month", "SUM(estimatedvalue) GROUP BY planned reco month", "msd data (field exists, NOT surfaced)", "P1 🔴", "✅ Ready"),
    ("Revenue Recognition", "Reco Walk (Won→Impl→Invoiced)", "Stage progress from Won date to SAP invoice date", "agfa_saporderid + SAP posting date", "P1 🔴", "✅ Ready"),
    ("Revenue Recognition", "Overdue Reco Alert", "Won deals where planned Reco date < today AND no SAP posting", "msd data + FeedFile", "P1 🔴", "✅ Ready"),
    ("Revenue Recognition", "Implementation Lag", "plannedrevenuerecognitiondate − estimatedclosedate (days)", "msd data — derivable", "P2 🟡", "✅ Ready"),
    ("Order Book / Backlog", "Order Book Value", "SUM(Won order value) WHERE not yet invoiced in SAP", "Requires EDW — agfa_saporderid → SAP order status", "P7 🔴", "🔴 EDW"),
    ("Order Book / Backlog", "Order Book by Equipment Type", "Same, split by product line", "EDW required", "P7 🔴", "🔴 EDW"),
    ("Order Book / Backlog", "Order Book by Region", "Same, split by region", "EDW required", "P7 🔴", "🔴 EDW"),
    ("Book & Bill", "B&B Volume", "Deals where OIT and Reco occur in same period", "agfa_requesteddeliverydate vs estimatedclosedate", "P3 🟡", "⚠ Partial"),
    ("Book & Bill", "B&B % of Total OIT", "B&B deals / Total OIT deals", "Derived from above", "P3 🟡", "⚠ Partial"),
    ("Win / Loss", "Win Rate % (Overall)", "Won / (Won + Lost)", "statecodename in opportunity.csv", "P3 🟡", "✅ Ready"),
    ("Win / Loss", "Win Rate by Region", "Won / (Won + Lost) per region", "opportunity.csv", "P3 🟡", "✅ Ready"),
    ("Win / Loss", "Win Rate by Equipment", "Won / (Won + Lost) per equipment type", "opportunity.csv", "P3 🟡", "✅ Ready"),
    ("Win / Loss", "Win Rate by Deal Size Band", "Won / (Won + Lost) per value band", "opportunity.csv", "P3 🟡", "✅ Ready"),
    ("Win / Loss", "Avg Time to Close (Won vs Lost)", "actualclosedate − createdon (days)", "opportunity.csv", "P3 🟡", "✅ Ready"),
    ("Win / Loss", "Loss Volume by Quarter", "COUNT(Lost) by quarter", "opportunity.csv", "P3 🟡", "✅ Ready"),
    ("2× Upside Rule", "Upside / Included with Risk Ratio", "Upside EUR ÷ Included with Risk EUR per region", "DataWeek.csv", "P2 ✅", "✅ Ready"),
    ("2× Upside Rule", "2× Rule Alert Flag", "Alert when ratio < 2 per region", "Derived from above", "P2 ✅", "✅ Ready"),
    ("Large Deal Tracker", "Large Deals List (>500 kEUR)", "Opportunities above threshold by stage/flag", "opportunity.csv", "P5 🟡", "✅ Ready"),
    ("Large Deal Tracker", "Stage Change Alert", "Flag deals where forecast category degraded week-on-week", "Weekly snapshots in Report 1", "P5 🟡", "⚠ Partial"),
    ("Services P&L", "Services Revenue", "Net Turnover TYPE=Support", "FeedFile.csv", "P6 🟡", "⚠ Partial"),
    ("Services P&L", "Services Margin %", "(Services Revenue − Services Cost) / Services Revenue", "FeedFile aggregate level only", "P6 🟡", "⚠ Estimated"),
    ("NWC / DSO", "Net Working Capital", "Receivables − Payables", "Not in any report — requires SAP FI extract", "P7 🔴", "🔴 EDW"),
    ("NWC / DSO", "DSO (Days Sales Outstanding)", "AR balance / Daily Revenue", "SAP FI data needed", "P7 🔴", "🔴 EDW"),
]

DH_EXISTING = [
    ("Revenue & Orders", "YTD Revenue", "SUM(amount) WHERE currency=EUR, status=2, year=2026", "orders.json", "EUR", "✅ Verified"),
    ("Revenue & Orders", "Full Year Forecast", "Actuals + Committed + Uncommitted + Unidentified (FY2026)", "kpis.json → forecast files", "EUR", "✅ Verified"),
    ("Revenue & Orders", "vs Budget %", "(Forecast − Budget) / Budget × 100", "BUD 2026 = €17.3M", "%", "✅ Verified"),
    ("Revenue & Orders", "vs Prior Year %", "(Forecast − LY) / LY × 100", "ACT LY from FY 2025.xls", "%", "✅ Verified"),
    ("Revenue & Orders", "Revenue by Product", "EUR split by UTP500 / UTP220 / UTP500+", "orders.json", "EUR", "✅ Verified"),
    ("Revenue & Orders", "Customer Concentration (Top 5/10)", "Top 5 / Top 10 as % of total EUR revenue", "orders.json", "%", "✅ Verified"),
    ("Revenue & Orders", "Open Orders Count", "COUNT WHERE status=1", "orders.json", "Count", "✅ Verified"),
    ("Revenue & Orders", "Open Orders Value", "SUM(amount) WHERE status=1", "orders.json", "EUR", "✅ Verified"),
    ("Revenue & Orders", "Average Order Value", "Total EUR / Count of completed orders", "orders.json", "EUR", "✅ Verified"),
    ("Revenue & Orders", "Revenue by Year", "Annual EUR totals 2023–2026", "orders.json", "EUR", "✅ Verified"),
    ("Pipeline & Conversion", "Quote Conversion Rate", "Orders (isOrdered=true) / Total Quotations", "quotations.json", "%", "✅ Verified"),
    ("Pipeline & Conversion", "Pipeline Value", "SUM(totalAmount) WHERE isOrdered=false", "quotations.json", "EUR", "✅ Verified"),
    ("Pipeline & Conversion", "Pipeline Count", "COUNT WHERE isOrdered=false", "quotations.json", "Count", "✅ Verified"),
    ("Pipeline & Conversion", "Conversion Rate by Year", "2023: 8.8%, 2024: 24.1%, 2025: 24.8%", "quotations.json", "%", "✅ Verified"),
    ("Pipeline & Conversion", "Conversion Rate by Product", "UTP220: 22.9%, UTP500+: 16.5%, UTP500: 16.2%", "quotations.json", "%", "✅ Verified"),
    ("Pipeline & Conversion", "Avg Quote Size", "Total amount / Quote count", "quotations.json", "EUR", "✅ Verified"),
    ("Pipeline & Conversion", "Ordered vs Non-Ordered Avg Size", "Ordered avg €20,856 vs Non-ordered avg €197,046", "quotations.json", "EUR", "✅ Verified"),
    ("Margin & Profitability", "Gross Margin %", "(Net Turnover − Standard Cost Total) / Net Turnover × 100", "revenue_summary.json / margin_data.json", "%", "✅ Verified"),
    ("Margin & Profitability", "Gross Margin (Absolute)", "Net Turnover − Standard Cost Total", "margin_data.json", "EUR", "✅ Verified"),
    ("Margin & Profitability", "Standard Cost per m²", "UTP500: €115.86, UTP220: €95.47, UTP500+: €115.86", "FY 2025.xls — Mapping Standard Costprices", "EUR/m²", "✅ Verified"),
    ("Margin & Profitability", "MSP 2026", "UTP220: €111.07/m², UTP500: €102.10/m²", "Revenue Overview in Feb2026 file", "EUR/m²", "✅ Verified"),
    ("Margin & Profitability", "Average Selling Price (EUR/m²)", "EUR Revenue / Total m² sold, by product and year", "orders.json", "EUR/m²", "✅ Verified"),
    ("Margin & Profitability", "Margin by Product", "GM% per UTP type", "margin_data.json", "%", "✅ Verified"),
    ("Forecast & Plans", "Forecast Composition", "Actuals (27%) + Committed (35%) + Uncommitted (13%) + Unidentified (25%)", "forecast.json", "%", "⚠ Estimated"),
    ("Forecast & Plans", "Budget Gap", "FY2026 Forecast – Budget = −€11.1M (−64.3%)", "forecast files", "EUR", "✅ Verified"),
    ("Forecast & Plans", "Forecast Revision History", "BUD €17.3M → RFC2 €31.2M → Current €6.2M", "forecast_revisions.json", "EUR", "✅ Verified"),
    ("Forecast & Plans", "Long-Term Plans", "2027: €38.6M, 2028: €51.4M, 2029: €69.9M", "long_term_plans.json", "EUR", "⚠ Estimated"),
    ("Forecast & Plans", "Scenario Analysis (Bear/Base/Bull)", "Base / Bear (×0.7) / Bull (×1.2)", "Calculated from sliders", "EUR", "⚠ Estimated"),
    ("Forecast & Plans", "TK Nucera NEOM Batches", "PO 32017233 — Batches #6–#25, ~€949K each", "ACTFY2025 TK Nucera sheet", "EUR", "✅ Verified"),
    ("Customer 360°", "Lifetime Revenue per Customer", "SUM(amount) WHERE customer=X AND status=2", "orders.json", "EUR", "✅ Verified"),
    ("Customer 360°", "Quotation Conversion per Customer", "Converted quotes / Total quotes per customer", "quotations.json", "%", "✅ Verified"),
    ("Customer 360°", "Product Mix per Customer", "Revenue by product type for selected customer", "orders.json", "%", "✅ Verified"),
]

DH_PROPOSED = [
    ("Operational", "Quote-to-Cash Cycle Time", "Order Date − Sent Date (days)", "quotations.json — G3: no Quotation ID in SAP", "P2 🟡", "⚠ Partial"),
    ("Operational", "On-Time Delivery Rate", "Delivered on/before Deliv. date", "AP1 SAP extract (Deliv. date vs Doc. Date)", "P3 🟡", "✅ Ready"),
    ("Operational", "Order Fulfillment Rate", "ConfirmQty / Order qty", "AP1 SAP extract", "P3 🟡", "✅ Ready"),
    ("Operational", "Avg Lead Time", "Deliv. date − Doc. Date (days)", "AP1 SAP extract", "P3 🟡", "✅ Ready"),
    ("Operational", "Pricing Variance", "Actual EUR/m² vs Standard pricing flag", "quotations.json + standard costs", "P3 🟡", "✅ Ready"),
    ("Customer Analytics", "DSO (Days Sales Outstanding)", "AR balance / Daily Revenue — based on payment terms", "Only 23.4% of amountPaid populated — G4 HIGH", "P2 🔴", "🔴 EDW"),
    ("Customer Analytics", "New vs Repeat Customers", "First order date analysis per customer", "orders.json", "P3 🟡", "✅ Ready"),
    ("Customer Analytics", "Customer Lifetime Value", "Cumulative revenue per SAP Customer ID", "orders.json", "P3 🟡", "✅ Ready"),
    ("Customer Analytics", "Revenue by Channel (IC vs Direct)", "Intercompany (AGFA Japan/Korea) vs Direct 3rd party", "orders.json + thirdPartyOrIco flag", "P2 🟡", "✅ Ready"),
    ("Customer Analytics", "Customer Segment Revenue", "OEM vs End-User vs Research vs Intercompany", "3rd P or ICO flag exists; further segmentation needed — G7", "P4 🟡", "⚠ Partial"),
    ("Win / Loss Intelligence", "Quote Loss Reasons", "Why 82% of quotes do not convert", "Not captured — no lost reason field — G9", "P4 🔴", "🔴 EDW"),
    ("Win / Loss Intelligence", "Sales Rep Performance", "Revenue attributed per sales rep", "Not captured — no sales rep field — G8", "P4 🔴", "🔴 EDW"),
    ("Win / Loss Intelligence", "Forecast Accuracy", "Forecast vs Actual for closed periods", "forecast_revisions.json — partially structured — G14", "P4 🟡", "⚠ Partial"),
]

DPS_EXISTING = [
    ("Revenue & Margin", "Net Revenue (3rd Party)", "Sum of Net Sales to third-party customers", "AMSP Contribution", "EUR", "✅ Verified"),
    ("Revenue & Margin", "AMSP Margin %", "AMSP Contribution / Net Sales 3rdP × 100", "AMSP Contribution", "%", "✅ Verified"),
    ("Revenue & Margin", "CO-PA Gross Margin %", "Gross Margin / Net TO × 100", "CO-PA GMPCOPA_1", "%", "✅ Verified"),
    ("Revenue & Margin", "EBIT", "Adjusted EBIT + Non-recurring items", "RECO Analysis", "kEUR", "✅ Verified"),
    ("Revenue & Margin", "Budget Variance %", "(Actual − Budget) / Budget × 100", "RECO vs BP1 budget", "%", "🔵 Derived"),
    ("Revenue & Margin", "Monthly Net Sales Trend", "EUR net sales per month Jan–Dec 2025", "AMSP Contribution", "EUR", "✅ Verified"),
    ("Revenue & Margin", "AMSP Margin by Budget Class/BU", "Margin % per product line (Anapurna, INCA, Jeti, etc.)", "AMSP Contribution", "%", "✅ Verified"),
    ("Revenue & Margin", "RECO Full P&L", "Revenue → Mfg Contribution → Gross Margin → SG&A → Adj EBIT → EBIT → Overall Result", "RECO Analysis (KRECO20)", "kEUR", "✅ Verified"),
    ("Revenue & Margin", "IC Elimination", "Intercompany flows (€270.6M to exclude from external revenue)", "RECO Analysis", "kEUR", "✅ Verified"),
    ("Revenue & Margin", "External Revenue", "Total Revenue minus IC elimination", "RECO — derived", "kEUR", "🔵 Derived"),
    ("Order Intake & Pipeline", "OIT Units", "Count of order intake records by period", "Order follow-up (manual Excel)", "Units", "✅ Verified"),
    ("Order Intake & Pipeline", "RR Units", "Count of invoiced/recognized orders by period", "Order follow-up", "Units", "✅ Verified"),
    ("Order Intake & Pipeline", "Backlog Units", "OIT cumulative − RR cumulative", "Order follow-up — derived", "Units", "🔵 Derived"),
    ("Order Intake & Pipeline", "Delayed Orders", "Count of orders with revised RR > original RR", "Delayed tracker tab", "Units", "✅ Verified"),
    ("Order Intake & Pipeline", "Pending RR (2026 YTD)", "46 units in backlog (end Feb 2026)", "Order follow-up", "Units", "✅ Verified"),
    ("Order Intake & Pipeline", "Cancellations", "Count of cancelled orders", "Order follow-up", "Units", "✅ Verified"),
    ("Equipment Family", "OIT by Family", "Units per family (Anapurna, Jeti, Oberon, Onset, Speedset)", "Order follow-up (VLOOKUP → Family col)", "Units", "✅ Verified"),
    ("Equipment Family", "RR by Family", "Invoiced units per family", "Order follow-up", "Units", "✅ Verified"),
    ("Equipment Family", "OIT vs RR Gap", "Units placed vs units invoiced per period", "Order follow-up — derived", "Units", "🔵 Derived"),
]

DPS_PROPOSED = [
    ("EUR-Value Metrics", "OIT EUR Value", "EUR value of new order intake per period", "Order intake tracks units only — needs SAP link", "P1 🔴", "🔴 EDW"),
    ("EUR-Value Metrics", "Pipeline EUR Value", "Forward pipeline EUR value", "CRM or SAP extract needed", "P1 🔴", "🔴 EDW"),
    ("EUR-Value Metrics", "RR EUR per Period", "Revenue recognised per month in EUR", "Cross-file join: Order follow-up → AMSP/RECO", "P1 🔴", "⚠ Partial"),
    ("EUR-Value Metrics", "HW vs Consumable Split (EUR)", "Revenue by Financial Application dimension", "BW query field needed", "P1 🔴", "🔴 EDW"),
    ("Deal-Level Margin", "Deal-Level Gross Margin", "Margin per order — no CRM connection for DPS", "CRM or order-level SAP data needed", "P1 🔴", "🔴 EDW"),
    ("Deal-Level Margin", "No AMSP Rate Coverage", "€64M net sales (33.6%) has no AMSP rate — margin unknown", "AMSP rate assignment in SAP", "P1 🔴", "🔴 EDW"),
    ("Order-to-Cash Cycle", "Order-to-Invoice Cycle Time", "Invoice Month − Order Month per order", "Order follow-up (derivable from formula cols)", "P2 🟡", "✅ Ready"),
    ("Order-to-Cash Cycle", "OIT-to-RR Lag by Family", "Months between OIT and RR per equipment family", "Order follow-up — derived", "P2 🟡", "✅ Ready"),
    ("Order-to-Cash Cycle", "Delayed Order Impact (EUR)", "Delayed units × estimated ASP", "Requires EUR dimension", "P3 🟡", "⚠ Partial"),
    ("Regional Breakdown (EUR)", "Revenue by Region (EUR)", "RECO and AMSP files are BU-level — no region dimension", "BW query with Region dimension needed", "P2 🔴", "🔴 EDW"),
    ("Regional Breakdown (EUR)", "OIT by Region (EUR)", "Order follow-up has Sales Org but not standard region codes", "Region mapping + EUR value needed", "P3 🟡", "⚠ Partial"),
]

HEIT_EXISTING = [
    ("OI — Order Intake", "MONTH ACT", "Monthly OI actual — new orders booked in the month", "oi_monthly.json (Key Figure = MONTH ACT)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "MONTH BUD", "Monthly OI budget — annual plan phased monthly", "oi_monthly.json (Key Figure = MONTH BUD)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "MONTH LY", "Monthly OI prior year — same calendar month, prior year", "oi_monthly.json (Key Figure = MONTH LY)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "MONTH FOR", "Monthly OI forecast — actuals YTD + forecast balance", "oi_monthly.json (Key Figure = MONTH FOR)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "YTD ACT", "Year-to-date cumulative OI actuals from Jan to reporting month", "oi_ytd.json (Key Figure = YTD ACT)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "YTD BUD", "Year-to-date OI budget cumulated to same period", "oi_ytd.json (Key Figure = YTD BUD)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "YTD LY", "Year-to-date OI prior year cumulated to same period", "oi_ytd.json (Key Figure = YTD LY)", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "FY (A+F)", "Full-year Actual + Forecast (YTD actuals + remaining forecast)", "oi_monthly.json (Key Figure = FY (A+F))", "kEUR", "✅ Verified"),
    ("OI — Order Intake", "FY BUD", "Full-year OI budget (static annual plan)", "oi_monthly.json (Key Figure = FY BUD)", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Planned Current Year", "Orders in backlog planned for recognition in current FY — Low risk", "ob_timeline.json / ob_regional.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Planned Next Years", "Backlog planned beyond current FY (multi-year) — Medium risk", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Rev Overdue ≤ 6 months", "Revenue planned in past period, not yet recognised — High risk", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Rev Overdue > 6 months", "Revenue overdue >6 months — Critical risk, escalation needed", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Not Planned Sales Order", "Sales orders received but not yet scheduled for recognition — Medium risk", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Not Planned Opportunity ≤ 3 months", "Near-term unplanned opportunity-stage orders", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Not Planned Opportunity 4–6 months", "Mid-range unplanned opportunities", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("OB — Order Book Buckets", "Not Planned Opportunity 7–24 months", "Longer-horizon unplanned opportunities", "ob_timeline.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 02 — Net Sales Hardware/Equipment", "Revenue from hardware equipment sales (net of returns/discounts)", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 07 — Net Sales Own Licenses", "Revenue from proprietary software licenses", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 09 — Net Sales subs. Own IP", "Subscription revenue for own IP software", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 11 — Net Sales 3rd Party Licenses", "Revenue from reselling third-party software licenses", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 26 — Net Sales Total", "Total revenue across all categories (sum of lines 02–25)", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 55 — TACO Margin", "Gross profit after direct costs. TACO Margin % = Line 55 / Line 26", "taco_key_lines.json", "kEUR / %", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 63 — Product Contribution", "TACO Margin minus product-specific overheads", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — P&L Key Lines", "Line 85 — TACO Contribution (EBIT)", "Operating result after all allocated expenses — KEY BOTTOM LINE", "taco_key_lines.json", "kEUR", "✅ Verified"),
    ("TACO — Computed Variants (14)", "ACT", "Actuals (EUR)", "taco_monthly / taco_by_month_bu", "kEUR", "✅ Verified"),
    ("TACO — Computed Variants (14)", "BUD", "Budget (EUR)", "taco_monthly", "kEUR", "✅ Verified"),
    ("TACO — Computed Variants (14)", "ACT LY", "Prior year actuals (EUR)", "taco_monthly", "kEUR", "✅ Verified"),
    ("TACO — Computed Variants (14)", "ACT vs BUD %", "(ACT − BUD) / BUD × 100", "Derived", "%", "🔵 Derived"),
    ("TACO — Computed Variants (14)", "ACT vs LY %", "(ACT − ACT LY) / ACT LY × 100", "Derived", "%", "🔵 Derived"),
    ("TACO — Computed Variants (14)", "ACT vs BUD (absolute)", "ACT − BUD", "Derived", "kEUR", "🔵 Derived"),
    ("TACO — Computed Variants (14)", "ACT vs LY (absolute)", "ACT − ACT LY", "Derived", "kEUR", "🔵 Derived"),
    ("TACO — Computed Variants (14)", "+ 7 Local Currency (LC) variants", "Same 7 metrics in local currency via selectable FX rate", "taco_monthly (LC columns)", "LC", "✅ Verified"),
    ("TACO — Computed Variants (14)", "ACT LC / BUD LC / LY LC + variances", "LC ACT, LC BUD, LC LY, LC vs BUD %, LC vs LY %, LC vs BUD abs, LC vs LY abs", "taco_monthly (LC columns)", "LC", "✅ Verified"),
    ("OI Business Type", "Net New", "New customer or entirely new product line", "oi_business_type.json", "kEUR", "✅ Verified"),
    ("OI Business Type", "Cross-sell", "Additional product to existing customer", "oi_business_type.json", "kEUR", "✅ Verified"),
    ("OI Business Type", "Upsell", "Upgraded version of existing product", "oi_business_type.json", "kEUR", "✅ Verified"),
    ("OI Business Type", "Renewal", "Contract renewal (same product, same customer)", "oi_business_type.json", "kEUR", "✅ Verified"),
    ("OI Business Type", "Upgrade and Updates", "Standard upgrades / maintenance contract renewals", "oi_business_type.json", "kEUR", "✅ Verified"),
]

HEIT_PROPOSED = [
    ("Revenue Lifecycle Funnel", "OI → OB → TACO Funnel", "OI booked → OB backlog → TACO recognised — end-to-end conversion", "3 separate Access DBs with no shared key — EDW required", "P7 🔴", "🔴 EDW"),
    ("Revenue Lifecycle Funnel", "OI-to-OB Conversion Rate", "% of OI that flows into confirmed Order Book", "No project/opportunity ID linking OI to OB", "P7 🔴", "🔴 EDW"),
    ("Revenue Lifecycle Funnel", "OB-to-TACO Recognition Rate", "% of Order Book released into TACO net sales in period", "No shared key between OB and TACO", "P7 🔴", "🔴 EDW"),
    ("Revenue Lifecycle Funnel", "Revenue Coverage %", "TACO Net Sales / OI Budget as %", "TACO 2026 budget not yet available", "P3 🟡", "⚠ Partial"),
    ("OI Enhancements", "MONTH FOR in P&L Report Charts", "FOR column already in data (13 snapshots) — not surfaced in UI", "oi_monthly.json — ready now", "P2 ✅", "✅ Ready"),
    ("OI Enhancements", "OI by Business Type at OB Level", "Business Type exists in OI but absent from OB files", "OB files have no Business Type field — EDW required", "P5 🔴", "🔴 EDW"),
    ("OI Enhancements", "Customer-Level OI", "OI aggregates have no customer dimension (only raw cache)", "Raw OI cache has customer — not extracted", "P5 🔴", "🔴 EDW"),
    ("OB Enhancements", "OB Age Analysis", "How long orders sit in each bucket before recognition", "ob_grid.json (pl_year+pl_qtr vs rep_year+rep_month) — derivable", "P2 🟡", "✅ Ready"),
    ("OB Enhancements", "Project Risk Level", "Risk flag per project-level order", "Not in any current source file", "P4 🔴", "🔴 EDW"),
    ("OB Enhancements", "Project Comments / Notes", "Qualitative commentary per project line", "Not in any current source file", "P4 🔴", "🔴 EDW"),
    ("OB Enhancements", "Rev Overdue Alert", "Auto-flag projects where overdue bucket crosses threshold", "ob_grid.json — derivable", "P2 🟡", "✅ Ready"),
    ("TACO Enhancements", "TACO FOR (Forecast Column)", "Forecast column in TACO P&L — available in OI but not TACO", "Access DB gap — TACO source lacks FOR series", "P4 🔴", "🔴 EDW"),
    ("TACO Enhancements", "TACO Quarterly Phasing", "Quarterly breakdown of P&L (currently monthly only)", "Monthly data — derivable by summing", "P2 🟡", "✅ Ready"),
    ("TACO Enhancements", "TACO by Company Code", "Entity-level P&L attribution (80+ Comp c values available)", "taco_monthly.json — partial", "P5 🟡", "⚠ Partial"),
    ("TACO Enhancements", "Services P&L Line Detail", "Breakdown of Implementation Services revenue/cost", "fa_line dimension in TACO — partial", "P5 🟡", "⚠ Partial"),
]

# ═══════════════════════════════════════════════════════════════════════════════
# COMBINED SHEET BUILDER
# Columns: #, Type, Section, KPI Name, Formula/Description,
#           Source/Data, Unit, Confidence/Status, Priority
# ═══════════════════════════════════════════════════════════════════════════════

NCOLS = 9

def build_combined_sheet(ws, title, subtitle, hdr_color, sec_color,
                         existing_data, proposed_data, n_existing, n_proposed):
    headers = ["#", "Type", "Section / Category", "KPI Name",
               "Formula / Description", "Source / Data Available",
               "Unit", "Confidence / Status", "Priority"]
    widths  = [4, 11, 22, 28, 44, 36, 10, 16, 12]
    col_colors = [hdr_color] * NCOLS

    sheet_header(ws, title, subtitle, hdr_color, NCOLS)
    col_header_row(ws, 3, headers, col_colors, widths)

    row = 4
    kpi_num = 0

    # ── EXISTING block ─────────────────────────────────────────────────────────
    divider_row(ws, row,
                f"  EXISTING KPIs  ({n_existing} total)",
                hdr_color, NCOLS)
    row += 1

    last_section = None
    for (section, kpi, formula, source, unit, conf) in existing_data:
        if section != last_section:
            section_row(ws, row, f"    {section}", sec_color, NCOLS)
            row += 1
            last_section = section
        kpi_num += 1
        ws.row_dimensions[row].height = 30

        # Type badge — Existing
        write_cell(ws, row, 1, kpi_num,   h_align="center")
        write_cell(ws, row, 2, "Existing", bold=True,
                   fcolor=C["exist_type"], tcolor=C["t_exist"], h_align="center")
        write_cell(ws, row, 3, section)
        write_cell(ws, row, 4, kpi, bold=True)
        write_cell(ws, row, 5, formula)
        write_cell(ws, row, 6, source)
        write_cell(ws, row, 7, unit, h_align="center")

        # Confidence styled
        cf_color, ct_color = CONF_STYLE.get(conf, (None, "000000"))
        write_cell(ws, row, 8, conf, fcolor=cf_color, tcolor=ct_color, h_align="center")

        # Priority — blank for existing
        write_cell(ws, row, 9, "", fcolor=C["light_grey"], h_align="center")
        row += 1

    # ── NEW / PROPOSED block ───────────────────────────────────────────────────
    ws.row_dimensions[row].height = 8  # small gap
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).fill = fill(C["white"])
    row += 1

    divider_row(ws, row,
                f"  NEW / PROPOSED KPIs  ({n_proposed} total)",
                hdr_color, NCOLS)
    row += 1

    last_section = None
    for (section, kpi, desc, data_avail, priority, status) in proposed_data:
        if section != last_section:
            section_row(ws, row, f"    {section}", sec_color, NCOLS)
            row += 1
            last_section = section
        kpi_num += 1
        ws.row_dimensions[row].height = 30

        write_cell(ws, row, 1, kpi_num,  h_align="center")
        write_cell(ws, row, 2, "New",    bold=True,
                   fcolor=C["new_type"], tcolor=C["t_new"], h_align="center")
        write_cell(ws, row, 3, section)
        write_cell(ws, row, 4, kpi, bold=True)
        write_cell(ws, row, 5, desc)
        write_cell(ws, row, 6, data_avail)
        write_cell(ws, row, 7, "",       fcolor=C["light_grey"], h_align="center")

        # Status styled
        sf_color, st_color = CONF_STYLE.get(status, (None, "000000"))
        write_cell(ws, row, 8, status, fcolor=sf_color, tcolor=st_color, h_align="center")

        # Priority styled
        pri_key = priority[:2] if len(priority) >= 2 else ""
        pf_color, pt_color = PRI_STYLE.get(pri_key, (C["light_grey"], "000000"))
        write_cell(ws, row, 9, priority, bold=True,
                   fcolor=pf_color, tcolor=pt_color, h_align="center")
        row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NCOLS)}3"
    return row

# ═══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── DR ─────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Digital Radiology")
build_combined_sheet(ws,
    "Digital Radiology — KPI Register",
    "30 Existing  |  23 New/Proposed  |  Sources: D365 CRM, SAP AP5/AP2/AP7, Sofon",
    C["hdr_dr"], C["sec_dr"], DR_EXISTING, DR_PROPOSED, 30, 23)

# ── DH ─────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Digital Hydrogen")
build_combined_sheet(ws,
    "Digital Hydrogen (Zirfon) — KPI Register",
    "32 Existing  |  13 New/Proposed  |  Sources: Sales Zirfon GHS, Quotation, AP1 SAP, FY2025.xls, Forecast",
    C["hdr_dh"], C["sec_dh"], DH_EXISTING, DH_PROPOSED, 32, 13)

# ── DPS ────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("DPS")
build_combined_sheet(ws,
    "DPS (Digital Print Solutions) — KPI Register",
    "19 Existing  |  11 New/Proposed  |  Sources: AMSP Contribution, CO-PA, RECO Analysis, Order Follow-up",
    C["hdr_dps"], C["sec_dps"], DPS_EXISTING, DPS_PROPOSED, 19, 11)

# ── HE IT ──────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("HE IT")
build_combined_sheet(ws,
    "HE IT — KPI Register",
    "39 Existing  |  15 New/Proposed  |  Sources: OI HEC pivot, OB Overview, OB Detailed, TACO (3 Access DBs)",
    C["hdr_heit"], C["sec_heit"], HEIT_EXISTING, HEIT_PROPOSED, 39, 15)

# ── SUMMARY ────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Summary", 0)
sheet_header(ws, "KPI Register — All Dashboards Summary",
    "Existing vs New KPIs across DR, DH, DPS, HE IT  |  Generated 2026-03-27",
    C["hdr_sum"], 8)

ws.row_dimensions[3].height = 10

count_hdrs = ["Dashboard", "Existing KPIs", "New / Proposed", "Total",
              "Ready to Build", "Need EDW", "Partial / Estimated", ""]
col_header_row(ws, 4, count_hdrs, [C["hdr_sum"]] * 8,
               [24, 16, 16, 12, 16, 14, 18, 4])

summary_data = [
    ("Digital Radiology (DR)", 30, 23, 53, 7, 8, 8),
    ("Digital Hydrogen (DH)", 32, 13, 45, 6, 3, 4),
    ("DPS", 19, 11, 30, 2, 6, 3),
    ("HE IT", 39, 15, 54, 4, 6, 5),
]
dash_colors = [C["hdr_dr"], C["hdr_dh"], C["hdr_dps"], C["hdr_heit"]]

for i, (dash, exist, prop, total, ready, edw, partial) in enumerate(summary_data, start=5):
    ws.row_dimensions[i].height = 22
    write_cell(ws, i, 1, dash,    bold=True,  fcolor=C["sec_sum"],   tcolor="1A2B4A",       h_align="left")
    write_cell(ws, i, 2, exist,   bold=True,  fcolor=C["verified"],  tcolor=C["t_verified"], h_align="center")
    write_cell(ws, i, 3, prop,    bold=True,  fcolor=C["new_type"],  tcolor=C["t_new"],      h_align="center")
    write_cell(ws, i, 4, total,   bold=True,  fcolor=C["sec_sum"],   tcolor="1A2B4A",        h_align="center")
    write_cell(ws, i, 5, ready,   fcolor=C["verified"],  tcolor=C["t_verified"],  h_align="center")
    write_cell(ws, i, 6, edw,     fcolor=C["proxy"],     tcolor=C["t_proxy"],     h_align="center")
    write_cell(ws, i, 7, partial, fcolor=C["estimated"], tcolor=C["t_estimated"], h_align="center")
    write_cell(ws, i, 8, "", fcolor=C["white"], border=False)

# Totals
ws.row_dimensions[9].height = 24
write_cell(ws, 9, 1, "TOTAL", bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 2, 120,  bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 3, 62,   bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 4, 182,  bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 5, 19,   bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 6, 23,   bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 7, 20,   bold=True, fcolor=C["hdr_sum"], tcolor=C["white"], h_align="center")
write_cell(ws, 9, 8, "", fcolor=C["white"], border=False)

ws.row_dimensions[10].height = 10

# Priority register
ws.row_dimensions[11].height = 16
ws.merge_cells("A11:H11")
sec_cell = ws.cell(row=11, column=1, value="  Priority Register — New / Proposed KPIs")
sec_cell.font = Font(bold=True, color="1A2B4A", size=11, name="Calibri")
sec_cell.fill = fill(C["sec_sum"])
sec_cell.alignment = align(h="left", v="center", wrap=False)
sec_cell.border = thin_border()

col_header_row(ws, 12, ["Priority", "Dashboard", "KPI", "Build Status", "Effort", "", "", ""],
               [C["hdr_sum"]] * 8)

priority_data = [
    ("P1 🔴", "DR",    "Revenue Recognition — Planned Reco by Month",     "✅ Ready",   "Very Low"),
    ("P1 🔴", "DR",    "Overdue Reco Alert",                               "✅ Ready",   "Very Low"),
    ("P1 🔴", "DPS",   "OIT EUR Value",                                    "🔴 EDW",     "High"),
    ("P1 🔴", "DPS",   "Pipeline EUR Value",                               "🔴 EDW",     "High"),
    ("P2 ✅", "DR",    "2× Upside Rule KPI + Alert",                       "✅ Ready",   "Very Low"),
    ("P2 🟡", "HE IT", "MONTH FOR in P&L Report Charts",                   "✅ Ready",   "Very Low"),
    ("P2 🟡", "HE IT", "OB Age Analysis / Rev Overdue Alert",              "✅ Ready",   "Low"),
    ("P2 🟡", "DH",    "Revenue by Channel (IC vs Direct)",                "✅ Ready",   "Low"),
    ("P3 🟡", "DR",    "Win / Loss Analysis (Win Rate, Avg Time to Close)", "✅ Ready",  "Low"),
    ("P3 🟡", "DR",    "Book & Bill Tracking",                             "⚠ Partial",  "Medium"),
    ("P3 🟡", "DH",    "On-Time Delivery Rate",                            "✅ Ready",   "Low"),
    ("P3 🟡", "DH",    "Order Fulfillment Rate & Avg Lead Time",           "✅ Ready",   "Low"),
    ("P4 🟡", "DH",    "New vs Repeat Customers",                          "✅ Ready",   "Low"),
    ("P5 🟡", "DR",    "Large Deal Tracker (>500 kEUR)",                   "✅ Ready",   "Medium"),
    ("P6 🟡", "DR",    "Reference Data Automation",                        "⚠ Partial",  "High"),
    ("P7 🔴", "DR",    "Order Book (Won not invoiced) — SAP data",         "🔴 EDW",     "High"),
    ("P7 🔴", "HE IT", "OI → OB → TACO Revenue Lifecycle Funnel",         "🔴 EDW",     "High"),
    ("P7 🔴", "HE IT", "TACO FOR (Forecast Column)",                       "🔴 EDW",     "Medium"),
    ("P7 🔴", "DR",    "NWC / DSO — SAP FI data",                         "🔴 EDW",     "High"),
]

for i, (pri, dash, kpi, status, effort) in enumerate(priority_data, start=13):
    ws.row_dimensions[i].height = 22
    pri_key = pri[:2]
    pfc, ptc = PRI_STYLE.get(pri_key, (C["light_grey"], "000000"))
    sf, st = CONF_STYLE.get(status, (C["light_grey"], "000000"))
    write_cell(ws, i, 1, pri,    bold=True,  fcolor=pfc,          tcolor=ptc,     h_align="center")
    write_cell(ws, i, 2, dash,   bold=True,  fcolor=C["sec_sum"], tcolor="1A2B4A",h_align="center")
    write_cell(ws, i, 3, kpi,    fcolor=None,                     tcolor="000000",h_align="left")
    write_cell(ws, i, 4, status, fcolor=sf,                       tcolor=st,      h_align="center")
    write_cell(ws, i, 5, effort, fcolor=C["light_grey"],          tcolor="000000",h_align="center")
    for c in [6, 7, 8]:
        write_cell(ws, i, c, "", fcolor=C["white"], border=False)

ws.freeze_panes = "A5"

# ── Save ───────────────────────────────────────────────────────────────────────
out = r"c:\Users\vajra\OneDrive\Documents\Work\AGFA Analysis\AgfaAnalytics\KPI_Register_All_Dashboards.xlsx"
wb.save(out)
print(f"Saved: {out}")
